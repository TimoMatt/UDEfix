import os
from itertools import chain, combinations

from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from fixpointtool.content.fpt_content import AccessDictionaries
from fixpointtool.content.fpt_mapping import FPTMapping
from fixpointtool.content.fpt_relation import FPTRelation
from nodeeditor.utils import dumpException


def getFileDialogFilter():
    return 'Graph (*.xlsx);;All files (*)'


def getFileDialogDirectory():
    return 'fixpointtool/content/excel'


def power_set(Y):
    s = list(Y)
    temp = chain.from_iterable(combinations(s, r) for r in range(len(s)+1))
    return frozenset(frozenset(i) for i in temp)


def cross_product(X, Y):
    return frozenset([(x, y) for x in X for y in Y])


class FPTImportDialog(QDialog):
    applications = ["Termination probability", "Stochastic game", "Bisimilarity"]

    def __init__(self, parent=None):
        super().__init__(parent)

        self.setWindowTitle("Import content")
        self.setWindowFlag(Qt.WindowContextHelpButtonHint, False)
        self.resize(300, 100)

        self.accessDicts = AccessDictionaries()

        self.filename = ""

        self.application_label = QLabel("Application:")
        self.importDirectory_label = QLabel("Import from:")

        self.application_comboBox = QComboBox()
        self.application_comboBox.addItems(FPTImportDialog.applications)

        self.importDirectory_button = QPushButton("...")
        self.importDirectory_button.setAutoDefault(False)
        self.importDirectory_button.setDefault(False)
        self.importDirectory_button.clicked.connect(self.onImportDirectoryButtonClicked)
        self.importDirectory_button.setObjectName("import_directory_button")

        self.application_hbox = QHBoxLayout()
        self.application_hbox.addWidget(self.application_label)
        self.application_hbox.addWidget(self.application_comboBox)

        self.importDirectory_hbox = QHBoxLayout()
        self.importDirectory_hbox.addWidget(self.importDirectory_label)
        self.importDirectory_hbox.addWidget(self.importDirectory_button)

        self.save_button = QPushButton("Save")
        self.cancel_button = QPushButton("Cancel")
        self.save_button.setDefault(False)
        self.save_button.setAutoDefault(False)
        self.cancel_button.setDefault(False)
        self.cancel_button.setAutoDefault(False)
        self.save_button.clicked.connect(self.onSave)
        self.cancel_button.clicked.connect(self.onCancel)

        self.save_cancel_box = QHBoxLayout()
        self.save_cancel_box.addWidget(self.save_button)
        self.save_cancel_box.addWidget(self.cancel_button)

        self.vbox = QVBoxLayout()
        self.vbox.addLayout(self.application_hbox)
        self.vbox.addLayout(self.importDirectory_hbox)
        self.vbox.addLayout(self.save_cancel_box)

        self.setLayout(self.vbox)

    def onImportDirectoryButtonClicked(self):
        fname, filter = QFileDialog.getOpenFileName(self, "Choose content directory", getFileDialogDirectory(), getFileDialogFilter())
        if fname != "" and os.path.isfile(fname):
            self.filename = fname
            self.importDirectory_button.setText(os.path.basename(fname))

    def onSave(self):
        if self.application_comboBox.currentText() == "Termination probability":
            try:
                states_name = "s"
                terminal_states_name = "t"
                mapping_name = "p"

                states = []
                terminal_states = []

                wb = load_workbook(self.filename)
                ws = wb.active

                rows = 0
                while ws['A' + str(rows + 1)].value is not None:
                    rows += 1

                for row in range(1, rows + 1):
                    colSum = 0
                    for col in range(1, rows + 1):
                        char = get_column_letter(col)
                        colSum += float(ws[char + str(row)].value)

                    if colSum == 0:
                        stateName = terminal_states_name + str(len(terminal_states))
                        states.append(stateName)
                        terminal_states.append(stateName)
                    elif colSum == 1:
                        stateName = states_name + str(len(states) - len(terminal_states))
                        states.append(stateName)
                    else:
                        print("WARNING: probability distribution did not sum up to 1")

                mappings = {}
                eta_mappings = []
                D = []

                for row in range(1, rows + 1):
                    if terminal_states_name in states[row - 1]:
                        continue
                    else:
                        listOfMappings = []
                        for col in range(1, rows + 1):
                            char = get_column_letter(col)
                            listOfMappings.append([states[col-1], float(ws[char + str(row)].value)])
                        newMapping = FPTMapping(mapping_name + str(len(mappings)), listOfMappings, 'miscellaneous', 'S', output_mv='algebra 1', output_mv_k='1')
                        mappings[mapping_name + str(len(mappings))] = newMapping
                        D.append(newMapping)
                        eta_mappings.append([states[row-1], newMapping])

                mappings['eta'] = FPTMapping('eta', eta_mappings, 'reindexing', 'S/T', 'D')
                mappings['k'] = FPTMapping('k', [[it, 1] for it in terminal_states], 'constant', 'T', output_mv='algebra 1', output_mv_k='1')

                self.accessDicts.updateDictionariesManually(mappings, {'S': frozenset(states), 'T': frozenset(terminal_states), 'S/T': frozenset([it for it in states if it not in terminal_states]), 'D': frozenset(D)}, {})

                self.accept()
            except Exception as e:
                msg = QMessageBox()
                msg.setWindowTitle("Warning")
                msg.setIcon(QMessageBox.Warning)
                msg.setText("Warning:\nError while importing graph")
                msg.exec()
        elif self.application_comboBox.currentText() == "Stochastic game":
            try:
                sink_states_name = "sink"
                min_states_name = "min"
                max_states_name = "max"
                av_states_name = "av"
                mapping_name = "p"

                states = []
                sink_states = []
                min_states = []
                max_states = []
                av_states = []

                wb = load_workbook(self.filename)
                ws = wb.active

                rows = 0
                while ws['A' + str(rows + 1)].value is not None:
                    rows += 1

                mappings = {}
                sets = {}
                relations = {}
                sink_values = []
                min_values = []
                max_values = []
                av_values = []

                for row in range(1, rows + 1):
                    rowType = ws['A' + str(row)].value

                    if rowType == "SINK":
                        stateName = sink_states_name + str(len(sink_states))
                        states.append(stateName)
                        sink_states.append(stateName)
                    elif rowType == "MIN":
                        stateName = min_states_name + str(len(min_states))
                        states.append(stateName)
                        min_states.append(stateName)
                    elif rowType == "MAX":
                        stateName = max_states_name + str(len(max_states))
                        states.append(stateName)
                        max_states.append(stateName)
                    elif rowType == "AV":
                        stateName = av_states_name + str(len(av_states))
                        states.append(stateName)
                        av_states.append(stateName)
                    else:
                        print("WARNING: Non-exisiting state type!")

                sets['Y'] = frozenset(states)
                sets['P_Y'] = power_set(states)
                sets['SINK'] = frozenset(sink_states)
                sets['MIN'] = frozenset(min_states)
                sets['MAX'] = frozenset(max_states)
                sets['AV'] = frozenset(av_states)

                for row in range(1, rows + 1):
                    rowType = ws['A' + str(row)].value

                    if rowType == "SINK":
                        sink_values.append(float(ws['B' + str(row)].value))
                    elif rowType == "MIN":
                        min_value = []
                        for col in range(2, rows + 2):
                            char = get_column_letter(col)
                            if ws[char + str(row)].value == 1:
                                min_value.append(states[col-2])
                        min_values.append(frozenset(min_value))
                    elif rowType == "MAX":
                        max_value = []
                        for col in range(2, rows + 2):
                            char = get_column_letter(col)
                            if ws[char + str(row)].value == 1:
                                max_value.append(states[col-2])
                        max_values.append(frozenset(max_value))
                    elif rowType == "AV":
                        listOfMappings = []
                        for col in range(2, rows + 2):
                            char = get_column_letter(col)
                            listOfMappings.append([states[col-2], float(ws[char + str(row)].value)])
                        newMapping = FPTMapping(mapping_name + str(len(mappings)), listOfMappings, 'miscellaneous', 'Y', output_mv='algebra 1', output_mv_k='1')
                        mappings[mapping_name + str(len(mappings))] = newMapping
                        av_values.append(newMapping)
                    else:
                        print("WARNING: Non-exisiting state type!")

                sets['D'] = frozenset(av_values)
                mappings['c1'] = FPTMapping('c1', [[sink_states[i], sink_values[i]] for i in range(len(sink_states))], 'constant', 'SINK', output_mv='algebra 1', output_mv_k='1')
                mappings['eta_min'] = FPTMapping('eta_min', [[min_states[i], min_values[i]] for i in range(len(min_states))], 'reindexing', 'MIN', 'P_Y')
                mappings['eta_max'] = FPTMapping('eta_max', [[max_states[i], max_values[i]] for i in range(len(max_states))], 'reindexing', 'MAX', 'P_Y')
                mappings['eta_av'] = FPTMapping('eta_av', [[av_states[i], av_values[i]] for i in range(len(av_states))], 'reindexing', 'AV', 'D')

                newRelation = []
                for inputElem in sets['Y']:
                    for outputElem in sets['P_Y']:
                        if inputElem in outputElem:
                            newRelation += [(inputElem, outputElem)]

                relations["R1"] = FPTRelation("R1", frozenset(newRelation), 'Y', 'P_Y', type="is-element-of")

                self.accessDicts.updateDictionariesManually(mappings, sets, relations)

                self.accept()
            except Exception as e:
                msg = QMessageBox()
                msg.setWindowTitle("Warning")
                msg.setIcon(QMessageBox.Warning)
                msg.setText("Warning:\nError while importing graph")
                msg.exec()
        elif self.application_comboBox.currentText() == "Bisimilarity":
            try:
                states_name = "s"

                states = []

                wb = load_workbook(self.filename)
                ws = wb.active

                rows = 0
                while ws['A' + str(rows + 1)].value is not None:
                    rows += 1

                mappings = {}
                sets = {}
                relations = {}

                for row in range(1, rows + 1):
                    stateName = states_name + str(len(states))
                    states.append(stateName)

                sets['X'] = frozenset(states)
                ps = power_set(states)
                sets['P_X'] = ps
                cp = cross_product(states, states)
                sets['XxX'] = cp
                sets['P_(XxX)'] = power_set(cp)
                sets['(P_X)x(P_X)'] = cross_product(ps, ps)

                successors = []

                for row in range(1, rows + 1):
                    row_successors = []

                    for col in range(1, rows + 1):
                        char = get_column_letter(col)
                        if str(ws[char + str(row)].value) == "1":
                            row_successors.append(states[col-1])

                    successors.append(frozenset(row_successors))

                listOfMappings = []
                for i in range(0, len(states)):
                    for j in range(0, len(states)):
                        listOfMappings.append([(states[i], states[j]), (successors[i], successors[j])])

                mappings['eta_bs'] = FPTMapping('eta_bs', listOfMappings, 'reindexing', 'XxX', '(P_X)x(P_X)')

                newRelation = []
                for inputElem in sets['XxX']:
                    for outputElem in sets['P_(XxX)']:
                        if inputElem in outputElem:
                            newRelation += [(inputElem, outputElem)]

                relations["R1"] = FPTRelation("R1", frozenset(newRelation), 'XxX', 'P_(XxX)', type="is-element-of")

                newRelation = []
                for inputElem in sets['P_(XxX)']:
                    x_values = []
                    y_values = []
                    for innerInputElem in inputElem:
                        x_values.append(innerInputElem[0])
                        y_values.append(innerInputElem[1])
                    for outputElem in sets['(P_X)x(P_X)']:
                        if set(x_values) == outputElem[0] and set(y_values) == outputElem[1]:
                            newRelation += [(inputElem, outputElem)]

                relations["R2"] = FPTRelation("R2", frozenset(newRelation), 'P_(XxX)', '(P_X)x(P_X)', type="projection", projection_type=2)

                self.accessDicts.updateDictionariesManually(mappings, sets, relations)

                self.accept()
            except Exception as e:
                dumpException(e)
                msg = QMessageBox()
                msg.setWindowTitle("Warning")
                msg.setIcon(QMessageBox.Warning)
                msg.setText("Warning:\nError while importing graph")
                msg.exec()

    def onCancel(self):
        self.reject()
