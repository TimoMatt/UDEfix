from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

if __name__ == '__main__':
    states_name = "s"
    terminal_states_name = "t"

    states = []
    terminal_states = []

    wb = load_workbook('excel/tp_graph.xlsx')
    ws = wb.active

    rows = 0
    while ws['A' + str(rows+1)].value is not None:
        rows += 1

    for row in range(1, rows+1):
        colSum = 0
        for col in range(1, rows+1):
            char = get_column_letter(col)
            colSum += float(ws[char + str(row)].value)

        if colSum == 0:
            stateName = terminal_states_name + str(len(terminal_states))
            states.append(stateName)
            terminal_states.append(stateName)
        elif colSum == 1:
            stateName = states_name + str(len(states) - len(terminal_states))
            states.append(stateName)
        else:
            print("WARNING: probability distribution did not sum up to 1")

    print(states)
    print(terminal_states)

    mappings = []

    for row in range(1, rows + 1):
        if terminal_states_name in states[row-1]:
            continue
        else:
            for col in range(1, rows + 1):
                char = get_column_letter(col)


